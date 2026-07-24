"""
Plot: Sept 6-7, 2023 hourly temperature (NSRDB, shifted to local time) vs.
our HEAT_FACTORS load-multiplier shape for the ieee14 heatwave scenario
(use_cases/ieee14/4batt_dcbus4_heatwave.py in the Phase 2 repo), plus a
non-heat-event day (Sept 16, 2023) for baseline comparison.

Sept 16 was picked by checking every day in September 2023 in this same
NSRDB sheet: the heat wave ran roughly Sept 1-10 (each day 101-109F max),
and temperatures only cooled off starting around Sept 16, which has the
lowest daily max (88.3F) and lowest daily average (81.6F) of the whole
month - the closest thing to a "normal" September day in this dataset.

Shows the heat-event days and HEAT_FACTORS are shaped consistently (both
peak in the afternoon rather than overnight), and that the heat event was a
real, meaningfully hotter day relative to a normal day in the same month -
which is the basis for shaping HEAT_FACTORS the way we did rather than
applying a flat multiplier across all 24 hours.

NSRDB's "Hour" column is recorded in UTC (Time Zone = 0 in the sheet's own
metadata); the sheet also records "Local Time Zone" = -6, so we shift by
-6 hours to align with local (Central) clock time before plotting. This is
a simple modulo shift, not a full date-boundary reconstruction, so treat
the alignment as illustrative rather than minute-exact.

Run from this folder:
    python plot_heatwave_scenario.py
"""

import os
import openpyxl
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(HERE, "Weather and Contingency Planning.xlsx")
LOCAL_TZ_OFFSET = -6  # matches the sheet's own "Local Time Zone" metadata

# Same 24 values as use_cases/ieee14/4batt_dcbus4_heatwave.py's HEAT_FACTORS
HEAT_FACTORS = [
    1.00, 1.00, 1.00, 1.00, 1.00, 1.00,
    1.01, 1.03, 1.05, 1.08, 1.11, 1.13,
    1.15, 1.15, 1.15, 1.13, 1.10, 1.07,
    1.04, 1.02, 1.00, 1.00, 1.00, 1.00,
]


def load_hourly_temps(day: int) -> list[float]:
    """Return 24 hourly temperatures (F) for Sept `day`, 2023, reindexed to
    local time (UTC hour u -> local hour (u + LOCAL_TZ_OFFSET) % 24)."""
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb["NRSDB"]
    utc_temps = [None] * 24
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        y, m, d, h, mi, tc, tf = row[0], row[1], row[2], row[3], row[4], row[5], row[6]
        if m == 9 and d == day:
            utc_temps[h] = tf
    if any(v is None for v in utc_temps):
        raise ValueError(f"Missing hourly readings for Sept {day}, 2023")
    local_temps = [None] * 24
    for utc_hour, temp in enumerate(utc_temps):
        local_hour = (utc_hour + LOCAL_TZ_OFFSET) % 24
        local_temps[local_hour] = temp
    return local_temps


def main():
    sept6 = load_hourly_temps(6)
    sept7 = load_hourly_temps(7)
    sept16 = load_hourly_temps(16)
    hours = list(range(1, 25))

    fig, ax1 = plt.subplots(figsize=(10, 6))

    ax1.plot(hours, sept6, color="#e05c3a", linewidth=2, marker="o", markersize=4,
              label="Sept 6, 2023 temp (F) - heat event")
    ax1.plot(hours, sept7, color="#d62728", linewidth=2, marker="s", markersize=4,
              label="Sept 7, 2023 temp (F) - heat event")
    ax1.plot(hours, sept16, color="#7f7f7f", linewidth=2, marker="d", markersize=4,
              linestyle=":", label="Sept 16, 2023 temp (F) - normal day baseline")
    ax1.set_xlabel("Hour (local time)")
    ax1.set_ylabel("Temperature (F)")
    ax1.set_xticks(hours)
    ax1.grid(axis="y", linestyle="--", alpha=0.35)

    ax2 = ax1.twinx()
    baseline_factors = [1.0] * 24  # 4batt_dcbus4.py (no HEAT_FACTORS) = no scaling, every hour
    ax2.plot(hours, baseline_factors, color="#7f7f7f", linewidth=2, linestyle=":",
              label="Baseline scenario load multiplier (no heat, flat 1.0)")
    ax2.plot(hours, HEAT_FACTORS, color="#1f77b4", linewidth=2.5, linestyle="--",
              marker="^", markersize=4, label="Heat scenario HEAT_FACTORS (load multiplier)")
    ax2.set_ylabel("Load multiplier")
    ax2.set_ylim(0.95, 1.20)

    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labels1 + labels2, loc="upper left", fontsize=9)

    plt.title("ERCOT Sept 6-7, 2023 heat event vs. a normal day (Sept 16): temperature vs. HEAT_FACTORS shape\n"
              "(NSRDB station near Austin, TX; temps shifted UTC -> local)")
    plt.tight_layout()

    out_path = os.path.join(HERE, "heatwave_scenario.png")
    plt.savefig(out_path, dpi=150)
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
